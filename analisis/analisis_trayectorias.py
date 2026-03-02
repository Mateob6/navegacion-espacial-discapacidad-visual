#!/usr/bin/env python3
"""
Comprehensive Step-by-Step Trajectory Analysis
for the Spatial Navigation study (Triangle Completion Task).

Analyzes waypoint data from 'Posiciones_cada_tres_pasos.xlsx'.
Produces:
  - fig9:  Reconstructed return-leg trajectories (Group x Modality)
  - fig10: Path efficiency analysis
  - fig11: Lateral deviation over normalized path progress
  - fig12: Heading consistency / heading error evolution
  - Summary statistics and ANOVAs
"""

import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib.lines import Line2D
import seaborn as sns
from scipy import stats
from pathlib import Path
import openpyxl

# ── Configuration ──────────────────────────────────────────────────
DATA_FILE = Path("/Users/mateo.belalcazar/Desktop/articulos/Navegación ciegos/datos/Posiciones_cada_tres_pasos.xlsx")
FIG_DIR = Path("/Users/mateo.belalcazar/Desktop/articulos/Navegación ciegos/figuras")
FIG_DIR.mkdir(exist_ok=True)

TILE_CM = 25.0  # each grid tile ~ 25 cm
ORIGIN = (5, 15)  # target return position (grid)

# Modality code mapping
MODALITY_MAP = {
    'a': 'Kinesthetic',
    't': 'Passive Touch',
    'p': 'Functional Interaction',
    'c': 'Control'
}
MODALITY_ORDER = ['Kinesthetic', 'Passive Touch', 'Functional Interaction', 'Control']
MODALITY_SHORT = {'Kinesthetic': 'KIN', 'Passive Touch': 'TOUCH',
                  'Functional Interaction': 'ACTION', 'Control': 'CTRL'}

# Group assignments from Rejilla agrupados
# Group 1 = DV (visual impairment), Group 2 = NDV (sighted blindfolded)
# PV-prefixed sheets = NDV, PN/PM-prefixed sheets = DV
DV_REJILLA = ['PN2', 'PN3', 'PN4', 'PN5', 'PN6', 'PN8', 'PN9', 'PN11',
              'PN12', 'PN15', 'PN16', 'PN17', 'PN19', 'PN20']
NDV_REJILLA = ['PN1', 'PN7', 'PN10', 'PN13', 'PN14', 'PN18',
               'PN21', 'PN22', 'PN23', 'PN24', 'PN25', 'PN26', 'PN27', 'PN28']

# Sheet name mapping (some have PV prefix instead of PN, some have spaces)
SHEET_REMAP = {
    'PN1': 'PV1', 'PN7': 'PV7', 'PN10': 'PV10',
    'PN13': 'PV13', 'PN14': 'PV14', 'PN18': 'PV18',
    'PN9': ' PN9', 'PN11': ' PN11', 'PN15': 'PM15'
}

# Style
plt.rcParams.update({
    'font.family': 'serif',
    'font.serif': ['Times New Roman', 'DejaVu Serif', 'serif'],
    'font.size': 11,
    'axes.titlesize': 13,
    'axes.labelsize': 12,
    'xtick.labelsize': 10,
    'ytick.labelsize': 10,
    'legend.fontsize': 10,
    'figure.dpi': 300,
    'savefig.dpi': 300,
    'savefig.bbox': 'tight',
    'axes.spines.top': False,
    'axes.spines.right': False,
})

GROUP_COLORS = {'DV': '#2166AC', 'NDV': '#B2182B'}
MODALITY_COLORS = {
    'Kinesthetic': '#E69F00',
    'Passive Touch': '#56B4E9',
    'Functional Interaction': '#009E73',
    'Control': '#CC79A7'
}


# ── Data Loading ───────────────────────────────────────────────────

def parse_trial_code(code):
    """Parse trial code like 'A1p' -> (form, size, modality)."""
    code = code.strip()
    form_char = code[0].upper()
    size_char = code[1]
    mod_char = code[2].lower()

    form = 'Acute' if form_char == 'A' else 'Obtuse'
    size = '2m' if size_char == '1' else '4m'
    modality = MODALITY_MAP.get(mod_char, mod_char)

    return form, size, modality


def get_start_position(size):
    """Return the starting position (last guided point) based on triangle size."""
    if size == '2m':
        return (5, 8)
    else:
        return (5, 1)


def load_all_trajectories():
    """Load trajectory data from all participant sheets."""
    wb = openpyxl.load_workbook(DATA_FILE, read_only=True, data_only=True)

    all_data = []

    for rejilla_code, group in [(c, 'DV') for c in DV_REJILLA] + [(c, 'NDV') for c in NDV_REJILLA]:
        sheet_name = SHEET_REMAP.get(rejilla_code, rejilla_code)

        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' (for {rejilla_code}) not found, skipping.")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=3, values_only=True))

        for row in rows:
            if row[1] is None:
                continue

            trial_code = str(row[1]).strip()
            if len(trial_code) < 3:
                continue

            try:
                form, size, modality = parse_trial_code(trial_code)
            except (IndexError, KeyError):
                continue

            # Extract waypoints: columns 2,3 (first x,y), 4,5 (second x,y), ...
            # up to columns 17,18 (8th pair); columns 18,19 are the ideal final position
            waypoints = []
            for j in range(0, 16, 2):  # pairs at indices 2-3, 4-5, ..., 16-17
                col_x = j + 2  # 0-indexed in tuple
                col_y = j + 3
                if col_x < len(row) and col_y < len(row):
                    x_val = row[col_x]
                    y_val = row[col_y]
                    if x_val is not None and y_val is not None:
                        try:
                            waypoints.append((float(x_val), float(y_val)))
                        except (ValueError, TypeError):
                            break
                    else:
                        break
                else:
                    break

            # Get E (position error) from column index 20 (col U)
            pos_error = None
            if len(row) > 20 and row[20] is not None:
                try:
                    pos_error = float(row[20])
                except (ValueError, TypeError):
                    pass

            if len(waypoints) >= 2:
                all_data.append({
                    'participant': rejilla_code,
                    'sheet': sheet_name,
                    'group': group,
                    'trial_code': trial_code,
                    'form': form,
                    'size': size,
                    'modality': modality,
                    'waypoints': waypoints,
                    'n_waypoints': len(waypoints),
                    'pos_error': pos_error,
                })

    wb.close()
    return all_data


# ── Trajectory Metrics ─────────────────────────────────────────────

def euclidean(p1, p2):
    return np.sqrt((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2)


def compute_trajectory_metrics(trial):
    """Compute all trajectory metrics for a single trial."""
    wps = trial['waypoints']
    wps_cm = [(x * TILE_CM, y * TILE_CM) for x, y in wps]
    origin_cm = (ORIGIN[0] * TILE_CM, ORIGIN[1] * TILE_CM)

    start = wps_cm[0]
    endpoint = wps_cm[-1]

    # Ideal path = straight line from start to origin
    ideal_length = euclidean(start, origin_cm)

    # Actual path length (sum of segments)
    actual_length = sum(euclidean(wps_cm[i], wps_cm[i+1]) for i in range(len(wps_cm)-1))

    # Path efficiency
    efficiency = ideal_length / actual_length if actual_length > 0 else np.nan

    # Position error at endpoint (distance from origin)
    final_error = euclidean(endpoint, origin_cm)

    # Lateral deviations from ideal line
    # Ideal direction vector
    dx_ideal = origin_cm[0] - start[0]
    dy_ideal = origin_cm[1] - start[1]
    line_len = np.sqrt(dx_ideal**2 + dy_ideal**2)

    lateral_devs = []
    progress_fracs = []
    heading_errors = []

    if line_len > 0:
        # Unit vector along ideal path
        ux = dx_ideal / line_len
        uy = dy_ideal / line_len

        # Ideal heading angle
        ideal_heading = np.degrees(np.arctan2(dx_ideal, dy_ideal))

        for i, pt in enumerate(wps_cm):
            # Vector from start to point
            vx = pt[0] - start[0]
            vy = pt[1] - start[1]

            # Projection along ideal line (progress)
            proj = vx * ux + vy * uy
            progress = proj / line_len if line_len > 0 else 0
            progress = np.clip(progress, 0, 2)  # allow overshoot

            # Perpendicular distance (lateral deviation)
            # Using cross product: |v x u|
            lat_dev = abs(vx * uy - vy * ux)

            lateral_devs.append(lat_dev)
            progress_fracs.append(progress)

            # Heading error at each waypoint (from waypoint i to i+1)
            if i < len(wps_cm) - 1:
                dx_move = wps_cm[i+1][0] - wps_cm[i][0]
                dy_move = wps_cm[i+1][1] - wps_cm[i][1]
                if abs(dx_move) + abs(dy_move) > 0:
                    # Heading toward origin from current point
                    dx_to_origin = origin_cm[0] - pt[0]
                    dy_to_origin = origin_cm[1] - pt[1]

                    actual_heading = np.degrees(np.arctan2(dx_move, dy_move))
                    ideal_heading_local = np.degrees(np.arctan2(dx_to_origin, dy_to_origin))

                    # Heading error (signed, wrapped to -180..180)
                    h_err = actual_heading - ideal_heading_local
                    h_err = ((h_err + 180) % 360) - 180
                    heading_errors.append(h_err)
                else:
                    heading_errors.append(np.nan)

    return {
        'ideal_length': ideal_length,
        'actual_length': actual_length,
        'efficiency': efficiency,
        'final_error': final_error,
        'lateral_devs': lateral_devs,
        'progress_fracs': progress_fracs,
        'heading_errors': heading_errors,
        'waypoints_cm': wps_cm,
        'origin_cm': origin_cm,
    }


# ── Main Analysis ──────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("TRAJECTORY ANALYSIS - Spatial Navigation Study")
    print("=" * 70)

    # Load data
    print("\nLoading trajectory data...")
    all_trials = load_all_trajectories()
    print(f"  Loaded {len(all_trials)} trials from {len(set(t['participant'] for t in all_trials))} participants")

    # Verify group assignment by checking position errors
    dv_errors = [t['pos_error'] for t in all_trials if t['group'] == 'DV' and t['pos_error'] is not None]
    ndv_errors = [t['pos_error'] for t in all_trials if t['group'] == 'NDV' and t['pos_error'] is not None]
    print(f"\n  Group verification (position error from E column):")
    print(f"    DV  mean E = {np.mean(dv_errors):.1f} cm (n={len(dv_errors)} trials)")
    print(f"    NDV mean E = {np.mean(ndv_errors):.1f} cm (n={len(ndv_errors)} trials)")
    if np.mean(dv_errors) < np.mean(ndv_errors):
        print("    >> CONFIRMED: DV has lower errors (better navigators)")
    else:
        print("    >> WARNING: Group assignment may be incorrect!")

    # Count by group and modality
    print("\n  Trials by Group x Modality:")
    for g in ['DV', 'NDV']:
        for m in MODALITY_ORDER:
            n = sum(1 for t in all_trials if t['group'] == g and t['modality'] == m)
            print(f"    {g} x {m}: {n}")

    # Compute metrics for each trial
    print("\nComputing trajectory metrics...")
    for trial in all_trials:
        metrics = compute_trajectory_metrics(trial)
        trial.update(metrics)

    # ── FIGURE 9: Trajectory Reconstruction ────────────────────────
    print("\n" + "=" * 70)
    print("FIGURE 9: Reconstructed Return-Leg Trajectories")
    print("=" * 70)

    fig, axes = plt.subplots(2, 4, figsize=(16, 9), sharex=False, sharey=False)

    for gi, group in enumerate(['DV', 'NDV']):
        for mi, modality in enumerate(MODALITY_ORDER):
            ax = axes[gi, mi]
            trials_subset = [t for t in all_trials if t['group'] == group and t['modality'] == modality]

            color = GROUP_COLORS[group]

            # Collect all trajectories for mean computation
            # Normalize to relative coordinates (start at 0,0)
            all_x_interp = []
            all_y_interp = []

            for trial in trials_subset:
                wps = trial['waypoints_cm']
                start = wps[0]

                # Relative coordinates
                xs = [p[0] - start[0] for p in wps]
                ys = [p[1] - start[1] for p in wps]

                # Plot individual trajectory (thin, semi-transparent)
                ax.plot(xs, ys, color=color, alpha=0.15, linewidth=0.7, zorder=1)
                ax.scatter(xs[-1], ys[-1], color=color, alpha=0.15, s=8, zorder=2)

                # Interpolate for mean trajectory
                if len(xs) >= 2:
                    # Cumulative distance as parameter
                    dists = [0]
                    for j in range(1, len(xs)):
                        dists.append(dists[-1] + np.sqrt((xs[j]-xs[j-1])**2 + (ys[j]-ys[j-1])**2))
                    total_dist = dists[-1]
                    if total_dist > 0:
                        t_norm = [d / total_dist for d in dists]
                        t_interp = np.linspace(0, 1, 20)
                        x_interp = np.interp(t_interp, t_norm, xs)
                        y_interp = np.interp(t_interp, t_norm, ys)
                        all_x_interp.append(x_interp)
                        all_y_interp.append(y_interp)

            # Plot mean trajectory
            if all_x_interp:
                mean_x = np.mean(all_x_interp, axis=0)
                mean_y = np.mean(all_y_interp, axis=0)
                ax.plot(mean_x, mean_y, color=color, linewidth=2.5, alpha=0.9, zorder=4)
                ax.scatter(mean_x[-1], mean_y[-1], color=color, s=40, zorder=5, edgecolors='white', linewidths=0.5)

            # Ideal line (from 0,0 to origin relative to average start)
            # Start positions vary by trial size, so compute average start and ideal endpoint
            starts_raw = [t['waypoints_cm'][0] for t in trials_subset]
            if starts_raw:
                avg_start = (np.mean([s[0] for s in starts_raw]), np.mean([s[1] for s in starts_raw]))
                origin_cm = (ORIGIN[0] * TILE_CM, ORIGIN[1] * TILE_CM)
                ideal_x = origin_cm[0] - avg_start[0]
                ideal_y = origin_cm[1] - avg_start[1]
                ax.plot([0, ideal_x], [0, ideal_y], 'k--', linewidth=1.2, alpha=0.5, zorder=3, label='Ideal')
                ax.scatter(ideal_x, ideal_y, marker='*', color='black', s=80, zorder=6, alpha=0.7)

            # Start point
            ax.scatter(0, 0, marker='s', color='black', s=40, zorder=6, alpha=0.7)

            ax.set_aspect('equal')
            if gi == 0:
                ax.set_title(modality, fontweight='bold')
            if mi == 0:
                ax.set_ylabel(f'{group}\nY displacement (cm)')
            if gi == 1:
                ax.set_xlabel('X displacement (cm)')

            ax.axhline(0, color='gray', linewidth=0.3, alpha=0.3)
            ax.axvline(0, color='gray', linewidth=0.3, alpha=0.3)

            n_t = len(trials_subset)
            n_p = len(set(t['participant'] for t in trials_subset))
            ax.text(0.02, 0.98, f'n={n_t} ({n_p}p)', transform=ax.transAxes,
                    fontsize=8, va='top', ha='left', color='gray')

    # Legend
    legend_elements = [
        Line2D([0], [0], color=GROUP_COLORS['DV'], linewidth=2.5, label='DV (mean)'),
        Line2D([0], [0], color=GROUP_COLORS['DV'], linewidth=0.7, alpha=0.3, label='DV (individual)'),
        Line2D([0], [0], color=GROUP_COLORS['NDV'], linewidth=2.5, label='NDV (mean)'),
        Line2D([0], [0], color=GROUP_COLORS['NDV'], linewidth=0.7, alpha=0.3, label='NDV (individual)'),
        Line2D([0], [0], color='black', linestyle='--', linewidth=1.2, alpha=0.5, label='Ideal path'),
        Line2D([0], [0], marker='*', color='black', linestyle='', markersize=8, alpha=0.7, label='Origin (target)'),
    ]
    fig.legend(handles=legend_elements, loc='lower center', ncol=6,
               bbox_to_anchor=(0.5, -0.02), frameon=False)

    fig.suptitle('Reconstructed Return-Leg Trajectories by Group and Modality',
                 fontweight='bold', fontsize=14, y=1.01)
    plt.tight_layout()
    fig.savefig(FIG_DIR / 'fig9_trajectories_reconstructed.png', dpi=300, bbox_inches='tight')
    plt.close()
    print("  Saved fig9_trajectories_reconstructed.png")


    # ── FIGURE 10: Path Efficiency ─────────────────────────────────
    print("\n" + "=" * 70)
    print("FIGURE 10: Path Efficiency Analysis")
    print("=" * 70)

    # Build dataframe for analysis
    df_metrics = pd.DataFrame([{
        'participant': t['participant'],
        'group': t['group'],
        'modality': t['modality'],
        'form': t['form'],
        'size': t['size'],
        'efficiency': t['efficiency'],
        'actual_length': t['actual_length'],
        'ideal_length': t['ideal_length'],
        'final_error': t['final_error'],
        'mean_lateral_dev': np.mean(t['lateral_devs']) if t['lateral_devs'] else np.nan,
        'max_lateral_dev': np.max(t['lateral_devs']) if t['lateral_devs'] else np.nan,
        'mean_heading_error': np.nanmean(np.abs(t['heading_errors'])) if t['heading_errors'] else np.nan,
    } for t in all_trials])

    # Cap efficiency at reasonable values (some near-zero actual paths)
    df_metrics['efficiency'] = df_metrics['efficiency'].clip(upper=2.0)

    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5))

    # Panel A: Violin plot
    ax = axes[0]
    plot_data = df_metrics.dropna(subset=['efficiency'])

    # Create positions for grouped violins
    positions_dv = [i - 0.2 for i in range(4)]
    positions_ndv = [i + 0.2 for i in range(4)]

    for gi, (group, positions) in enumerate(zip(['DV', 'NDV'], [positions_dv, positions_ndv])):
        color = GROUP_COLORS[group]
        for mi, modality in enumerate(MODALITY_ORDER):
            subset = plot_data[(plot_data['group'] == group) & (plot_data['modality'] == modality)]['efficiency']
            if len(subset) > 0:
                parts = ax.violinplot(subset, positions=[positions[mi]], showmeans=True, showmedians=False, widths=0.35)
                for pc in parts['bodies']:
                    pc.set_facecolor(color)
                    pc.set_alpha(0.3)
                    pc.set_edgecolor(color)
                for partname in ('cmeans', 'cmins', 'cmaxes', 'cbars'):
                    if partname in parts:
                        parts[partname].set_edgecolor(color)
                        parts[partname].set_linewidth(1.5)

                # Overlay individual points with jitter
                jitter = np.random.normal(0, 0.04, len(subset))
                ax.scatter(positions[mi] + jitter, subset, color=color, alpha=0.3, s=12, zorder=3)

    ax.set_xticks(range(4))
    ax.set_xticklabels([MODALITY_SHORT[m] for m in MODALITY_ORDER], fontsize=10)
    ax.set_ylabel('Path Efficiency (ideal / actual)')
    ax.set_xlabel('Sensory Modality')
    ax.axhline(1.0, color='gray', linestyle=':', linewidth=0.8, alpha=0.5)
    ax.set_title('A. Path Efficiency Distribution', fontweight='bold')

    legend_elements = [
        plt.Rectangle((0,0), 1, 1, facecolor=GROUP_COLORS['DV'], alpha=0.4, label='DV'),
        plt.Rectangle((0,0), 1, 1, facecolor=GROUP_COLORS['NDV'], alpha=0.4, label='NDV'),
    ]
    ax.legend(handles=legend_elements, loc='upper right', frameon=True, framealpha=0.9)

    # Panel B: Bar chart with error bars (mean +/- SE)
    ax = axes[1]

    bar_width = 0.35
    x_pos = np.arange(4)

    for gi, group in enumerate(['DV', 'NDV']):
        means = []
        sems = []
        for modality in MODALITY_ORDER:
            subset = plot_data[(plot_data['group'] == group) & (plot_data['modality'] == modality)]['efficiency']
            means.append(subset.mean())
            sems.append(subset.sem())

        offset = -bar_width/2 if gi == 0 else bar_width/2
        bars = ax.bar(x_pos + offset, means, bar_width, yerr=sems,
                      label=group, color=GROUP_COLORS[group], alpha=0.7,
                      edgecolor='white', linewidth=0.5,
                      capsize=4, error_kw={'linewidth': 1.2})

    ax.set_xticks(x_pos)
    ax.set_xticklabels([MODALITY_SHORT[m] for m in MODALITY_ORDER], fontsize=10)
    ax.set_ylabel('Mean Path Efficiency (+/- SE)')
    ax.set_xlabel('Sensory Modality')
    ax.axhline(1.0, color='gray', linestyle=':', linewidth=0.8, alpha=0.5)
    ax.set_title('B. Mean Path Efficiency by Group and Modality', fontweight='bold')
    ax.legend(loc='upper right', frameon=True, framealpha=0.9)

    plt.tight_layout()
    fig.savefig(FIG_DIR / 'fig10_path_efficiency.png', dpi=300, bbox_inches='tight')
    plt.close()
    print("  Saved fig10_path_efficiency.png")

    # ANOVA on path efficiency
    print("\n  --- 2(Group) x 4(Modality) ANOVA on Path Efficiency ---")

    # Aggregate to participant means for proper repeated-measures analysis
    df_agg = df_metrics.groupby(['participant', 'group', 'modality'])['efficiency'].mean().reset_index()

    # Mixed ANOVA: Group (between) x Modality (within)
    # Use Type III SS with manual computation
    # First, simple two-way ANOVA on aggregated data
    from itertools import product

    groups_list = ['DV', 'NDV']
    modalities_list = MODALITY_ORDER

    # Collect data for ANOVA
    anova_data = []
    for part in df_agg['participant'].unique():
        part_data = df_agg[df_agg['participant'] == part]
        group = part_data['group'].iloc[0]
        for mod in modalities_list:
            val = part_data[part_data['modality'] == mod]['efficiency']
            if len(val) > 0:
                anova_data.append({
                    'participant': part,
                    'group': group,
                    'modality': mod,
                    'efficiency': val.values[0]
                })

    df_anova = pd.DataFrame(anova_data)

    # Use pingouin if available, otherwise statsmodels
    try:
        import pingouin as pg
        aov = pg.mixed_anova(data=df_anova, dv='efficiency', within='modality',
                             between='group', subject='participant')
        print(aov.to_string(index=False))

        # Post-hoc
        print("\n  Post-hoc pairwise comparisons (Modality, Bonferroni):")
        posthoc = pg.pairwise_tests(data=df_anova, dv='efficiency', within='modality',
                                     between='group', subject='participant',
                                     padjust='bonf', parametric=True)
        print(posthoc.to_string(index=False))

    except ImportError:
        print("  (pingouin not available, using manual F-tests)")

        # Main effect of Group
        dv_eff = df_anova[df_anova['group'] == 'DV']['efficiency']
        ndv_eff = df_anova[df_anova['group'] == 'NDV']['efficiency']
        f_group, p_group = stats.f_oneway(dv_eff, ndv_eff)
        print(f"  Group effect: F = {f_group:.3f}, p = {p_group:.4f}")
        print(f"    DV mean = {dv_eff.mean():.3f} (SD={dv_eff.std():.3f})")
        print(f"    NDV mean = {ndv_eff.mean():.3f} (SD={ndv_eff.std():.3f})")

        # Main effect of Modality
        mod_groups = [df_anova[df_anova['modality'] == m]['efficiency'].values for m in modalities_list]
        f_mod, p_mod = stats.f_oneway(*mod_groups)
        print(f"\n  Modality effect: F = {f_mod:.3f}, p = {p_mod:.4f}")
        for m in modalities_list:
            subset = df_anova[df_anova['modality'] == m]['efficiency']
            print(f"    {m}: mean = {subset.mean():.3f} (SD={subset.std():.3f})")

        # Interaction (simple test: 2-way ANOVA via statsmodels)
        try:
            import statsmodels.api as sm
            from statsmodels.formula.api import ols

            model = ols('efficiency ~ C(group) * C(modality)', data=df_anova).fit()
            anova_table = sm.stats.anova_lm(model, typ=2)
            print(f"\n  Two-way ANOVA (Type II):")
            print(anova_table.to_string())
        except ImportError:
            print("  (statsmodels not available for interaction test)")

        # Pairwise t-tests for modality with Bonferroni
        print("\n  Pairwise t-tests (Modality, Bonferroni):")
        from itertools import combinations
        pairs = list(combinations(modalities_list, 2))
        n_comparisons = len(pairs)
        for m1, m2 in pairs:
            d1 = df_anova[df_anova['modality'] == m1]['efficiency']
            d2 = df_anova[df_anova['modality'] == m2]['efficiency']
            t_stat, p_val = stats.ttest_ind(d1, d2)
            p_adj = min(p_val * n_comparisons, 1.0)
            sig = '*' if p_adj < 0.05 else ''
            print(f"    {MODALITY_SHORT[m1]} vs {MODALITY_SHORT[m2]}: t={t_stat:.3f}, p={p_val:.4f}, p_adj={p_adj:.4f} {sig}")


    # ── FIGURE 11: Lateral Deviation ───────────────────────────────
    print("\n" + "=" * 70)
    print("FIGURE 11: Lateral Deviation Analysis")
    print("=" * 70)

    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5))

    # Panel A: Lateral deviation over normalized path progress by Group x Modality
    ax = axes[0]

    n_bins = 10
    bin_edges = np.linspace(0, 1, n_bins + 1)
    bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2

    linestyles = {'DV': '-', 'NDV': '--'}

    for group in ['DV', 'NDV']:
        for modality in MODALITY_ORDER:
            trials_subset = [t for t in all_trials if t['group'] == group and t['modality'] == modality]

            # Bin lateral deviations by progress fraction
            binned_devs = [[] for _ in range(n_bins)]
            for trial in trials_subset:
                for prog, dev in zip(trial['progress_fracs'], trial['lateral_devs']):
                    bin_idx = np.searchsorted(bin_edges[1:], prog, side='right')
                    bin_idx = min(bin_idx, n_bins - 1)
                    binned_devs[bin_idx].append(dev)

            means = [np.mean(b) if b else np.nan for b in binned_devs]
            sems = [np.std(b)/np.sqrt(len(b)) if len(b) > 1 else 0 for b in binned_devs]

            color = MODALITY_COLORS[modality]
            ls = linestyles[group]
            marker = 'o' if group == 'DV' else 's'

            valid = [i for i in range(n_bins) if not np.isnan(means[i])]
            if valid:
                x_plot = [bin_centers[i] * 100 for i in valid]
                y_plot = [means[i] for i in valid]
                yerr_plot = [sems[i] for i in valid]

                ax.errorbar(x_plot, y_plot, yerr=yerr_plot,
                           color=color, linestyle=ls, marker=marker, markersize=4,
                           linewidth=1.5, alpha=0.8, capsize=2, label=f'{group}-{MODALITY_SHORT[modality]}')

    ax.set_xlabel('Normalized Path Progress (%)')
    ax.set_ylabel('Lateral Deviation (cm)')
    ax.set_title('A. Lateral Deviation Over Path Progress', fontweight='bold')
    ax.legend(fontsize=7, ncol=2, loc='upper left', frameon=True, framealpha=0.9)

    # Panel B: Mean lateral deviation bar chart
    ax = axes[1]

    bar_width = 0.35
    x_pos = np.arange(4)

    for gi, group in enumerate(['DV', 'NDV']):
        means = []
        sems = []
        for modality in MODALITY_ORDER:
            subset = df_metrics[(df_metrics['group'] == group) & (df_metrics['modality'] == modality)]['mean_lateral_dev'].dropna()
            means.append(subset.mean())
            sems.append(subset.sem())

        offset = -bar_width/2 if gi == 0 else bar_width/2
        ax.bar(x_pos + offset, means, bar_width, yerr=sems,
               label=group, color=GROUP_COLORS[group], alpha=0.7,
               edgecolor='white', linewidth=0.5,
               capsize=4, error_kw={'linewidth': 1.2})

    ax.set_xticks(x_pos)
    ax.set_xticklabels([MODALITY_SHORT[m] for m in MODALITY_ORDER], fontsize=10)
    ax.set_ylabel('Mean Lateral Deviation (cm)')
    ax.set_xlabel('Sensory Modality')
    ax.set_title('B. Mean Lateral Deviation by Group and Modality', fontweight='bold')
    ax.legend(loc='upper right', frameon=True, framealpha=0.9)

    plt.tight_layout()
    fig.savefig(FIG_DIR / 'fig11_lateral_deviation.png', dpi=300, bbox_inches='tight')
    plt.close()
    print("  Saved fig11_lateral_deviation.png")

    # ANOVA on mean lateral deviation
    print("\n  --- 2(Group) x 4(Modality) ANOVA on Mean Lateral Deviation ---")
    df_agg_lat = df_metrics.groupby(['participant', 'group', 'modality'])['mean_lateral_dev'].mean().reset_index()
    df_agg_lat = df_agg_lat.dropna()

    # Filter for complete cases (participants with all 4 modalities)
    complete_lat = df_agg_lat.groupby('participant').filter(lambda x: len(x) == 4)
    n_complete_lat = complete_lat['participant'].nunique()
    print(f"  Complete cases: {n_complete_lat} participants (of {df_agg_lat['participant'].nunique()})")

    try:
        import pingouin as pg
        aov_lat = pg.mixed_anova(data=complete_lat, dv='mean_lateral_dev', within='modality',
                                  between='group', subject='participant')
        print(aov_lat.to_string(index=False))

        print("\n  Post-hoc (Group x Modality):")
        posthoc_lat = pg.pairwise_tests(data=complete_lat, dv='mean_lateral_dev', within='modality',
                                         between='group', subject='participant',
                                         padjust='bonf', parametric=True)
        # Print only group comparison and interaction rows
        print(posthoc_lat.to_string(index=False))
    except ImportError:
        dv_lat = complete_lat[complete_lat['group'] == 'DV']['mean_lateral_dev']
        ndv_lat = complete_lat[complete_lat['group'] == 'NDV']['mean_lateral_dev']
        f_group, p_group = stats.f_oneway(dv_lat, ndv_lat)
        print(f"  Group effect: F = {f_group:.3f}, p = {p_group:.4f}")
        print(f"    DV mean = {dv_lat.mean():.1f} cm (SD={dv_lat.std():.1f})")
        print(f"    NDV mean = {ndv_lat.mean():.1f} cm (SD={ndv_lat.std():.1f})")

        mod_groups = [complete_lat[complete_lat['modality'] == m]['mean_lateral_dev'].values for m in modalities_list]
        f_mod, p_mod = stats.f_oneway(*mod_groups)
        print(f"  Modality effect: F = {f_mod:.3f}, p = {p_mod:.4f}")

        try:
            import statsmodels.api as sm
            from statsmodels.formula.api import ols
            model_lat = ols('mean_lateral_dev ~ C(group) * C(modality)', data=complete_lat).fit()
            anova_lat_table = sm.stats.anova_lm(model_lat, typ=2)
            print(f"\n  Two-way ANOVA (Type II):")
            print(anova_lat_table.to_string())
        except ImportError:
            pass


    # ── FIGURE 12: Heading Consistency ─────────────────────────────
    print("\n" + "=" * 70)
    print("FIGURE 12: Heading Consistency Analysis")
    print("=" * 70)

    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5))

    # Panel A: Heading error evolution over waypoints
    ax = axes[0]

    max_waypoints = 7  # maximum number of heading segments

    for group in ['DV', 'NDV']:
        for modality in MODALITY_ORDER:
            trials_subset = [t for t in all_trials if t['group'] == group and t['modality'] == modality]

            # Collect heading errors by waypoint index
            heading_by_wp = [[] for _ in range(max_waypoints)]
            for trial in trials_subset:
                for wi, he in enumerate(trial['heading_errors']):
                    if wi < max_waypoints and not np.isnan(he):
                        heading_by_wp[wi].append(abs(he))

            # Plot means with error bars
            wp_indices = []
            means = []
            sems = []
            for wi in range(max_waypoints):
                if len(heading_by_wp[wi]) >= 3:
                    wp_indices.append(wi + 1)
                    means.append(np.mean(heading_by_wp[wi]))
                    sems.append(np.std(heading_by_wp[wi]) / np.sqrt(len(heading_by_wp[wi])))

            if wp_indices:
                color = MODALITY_COLORS[modality]
                ls = linestyles[group]
                marker = 'o' if group == 'DV' else 's'

                ax.errorbar(wp_indices, means, yerr=sems,
                           color=color, linestyle=ls, marker=marker, markersize=4,
                           linewidth=1.5, alpha=0.8, capsize=2,
                           label=f'{group}-{MODALITY_SHORT[modality]}')

    ax.set_xlabel('Waypoint Segment')
    ax.set_ylabel('Absolute Heading Error (degrees)')
    ax.set_title('A. Heading Error Evolution Over Trajectory', fontweight='bold')
    ax.legend(fontsize=7, ncol=2, loc='upper right', frameon=True, framealpha=0.9)
    ax.axhline(0, color='gray', linestyle=':', linewidth=0.5, alpha=0.5)

    # Panel B: Mean heading error bar chart
    ax = axes[1]

    bar_width = 0.35
    x_pos = np.arange(4)

    for gi, group in enumerate(['DV', 'NDV']):
        means = []
        sems = []
        for modality in MODALITY_ORDER:
            subset = df_metrics[(df_metrics['group'] == group) & (df_metrics['modality'] == modality)]['mean_heading_error'].dropna()
            means.append(subset.mean())
            sems.append(subset.sem())

        offset = -bar_width/2 if gi == 0 else bar_width/2
        ax.bar(x_pos + offset, means, bar_width, yerr=sems,
               label=group, color=GROUP_COLORS[group], alpha=0.7,
               edgecolor='white', linewidth=0.5,
               capsize=4, error_kw={'linewidth': 1.2})

    ax.set_xticks(x_pos)
    ax.set_xticklabels([MODALITY_SHORT[m] for m in MODALITY_ORDER], fontsize=10)
    ax.set_ylabel('Mean Absolute Heading Error (degrees)')
    ax.set_xlabel('Sensory Modality')
    ax.set_title('B. Mean Heading Error by Group and Modality', fontweight='bold')
    ax.legend(loc='upper right', frameon=True, framealpha=0.9)

    plt.tight_layout()
    fig.savefig(FIG_DIR / 'fig12_heading_consistency.png', dpi=300, bbox_inches='tight')
    plt.close()
    print("  Saved fig12_heading_consistency.png")

    # ANOVA on heading error
    print("\n  --- 2(Group) x 4(Modality) ANOVA on Mean Heading Error ---")
    df_agg_head = df_metrics.groupby(['participant', 'group', 'modality'])['mean_heading_error'].mean().reset_index()
    df_agg_head = df_agg_head.dropna()

    # Filter for complete cases
    complete_head = df_agg_head.groupby('participant').filter(lambda x: len(x) == 4)
    n_complete_head = complete_head['participant'].nunique()
    print(f"  Complete cases: {n_complete_head} participants (of {df_agg_head['participant'].nunique()})")

    try:
        import pingouin as pg
        aov_head = pg.mixed_anova(data=complete_head, dv='mean_heading_error', within='modality',
                                   between='group', subject='participant')
        print(aov_head.to_string(index=False))

        print("\n  Post-hoc (Group x Modality):")
        posthoc_head = pg.pairwise_tests(data=complete_head, dv='mean_heading_error', within='modality',
                                          between='group', subject='participant',
                                          padjust='bonf', parametric=True)
        print(posthoc_head.to_string(index=False))
    except ImportError:
        dv_head = complete_head[complete_head['group'] == 'DV']['mean_heading_error']
        ndv_head = complete_head[complete_head['group'] == 'NDV']['mean_heading_error']
        f_group, p_group = stats.f_oneway(dv_head, ndv_head)
        print(f"  Group effect: F = {f_group:.3f}, p = {p_group:.4f}")

        mod_groups = [complete_head[complete_head['modality'] == m]['mean_heading_error'].values for m in modalities_list]
        f_mod, p_mod = stats.f_oneway(*mod_groups)
        print(f"  Modality effect: F = {f_mod:.3f}, p = {p_mod:.4f}")

        try:
            import statsmodels.api as sm
            from statsmodels.formula.api import ols
            model_head = ols('mean_heading_error ~ C(group) * C(modality)', data=complete_head).fit()
            anova_head_table = sm.stats.anova_lm(model_head, typ=2)
            print(f"\n  Two-way ANOVA (Type II):")
            print(anova_head_table.to_string())
        except ImportError:
            pass


    # ── SUMMARY TABLE ──────────────────────────────────────────────
    print("\n" + "=" * 70)
    print("SUMMARY TABLE: All Metrics by Group x Modality")
    print("=" * 70)

    summary_rows = []
    for group in ['DV', 'NDV']:
        for modality in MODALITY_ORDER:
            subset = df_metrics[(df_metrics['group'] == group) & (df_metrics['modality'] == modality)]
            n_trials = len(subset)
            n_parts = subset['participant'].nunique()

            eff = subset['efficiency'].dropna()
            lat = subset['mean_lateral_dev'].dropna()
            head = subset['mean_heading_error'].dropna()
            ferr = subset['final_error'].dropna()

            summary_rows.append({
                'Group': group,
                'Modality': MODALITY_SHORT[modality],
                'N_trials': n_trials,
                'N_participants': n_parts,
                'Efficiency_M': f"{eff.mean():.3f}" if len(eff) > 0 else '-',
                'Efficiency_SD': f"{eff.std():.3f}" if len(eff) > 0 else '-',
                'LatDev_M(cm)': f"{lat.mean():.1f}" if len(lat) > 0 else '-',
                'LatDev_SD': f"{lat.std():.1f}" if len(lat) > 0 else '-',
                'HeadErr_M(deg)': f"{head.mean():.1f}" if len(head) > 0 else '-',
                'HeadErr_SD': f"{head.std():.1f}" if len(head) > 0 else '-',
                'FinalErr_M(cm)': f"{ferr.mean():.1f}" if len(ferr) > 0 else '-',
                'FinalErr_SD': f"{ferr.std():.1f}" if len(ferr) > 0 else '-',
            })

    df_summary = pd.DataFrame(summary_rows)
    print(df_summary.to_string(index=False))

    # Overall group means
    print("\n  Overall Group Means:")
    for group in ['DV', 'NDV']:
        subset = df_metrics[df_metrics['group'] == group]
        print(f"  {group}:")
        print(f"    Efficiency: {subset['efficiency'].dropna().mean():.3f} (SD={subset['efficiency'].dropna().std():.3f})")
        print(f"    Lateral Dev: {subset['mean_lateral_dev'].dropna().mean():.1f} cm (SD={subset['mean_lateral_dev'].dropna().std():.1f})")
        print(f"    Heading Err: {subset['mean_heading_error'].dropna().mean():.1f} deg (SD={subset['mean_heading_error'].dropna().std():.1f})")
        print(f"    Final Error: {subset['final_error'].dropna().mean():.1f} cm (SD={subset['final_error'].dropna().std():.1f})")

    # ── Additional effect sizes ────────────────────────────────────
    print("\n" + "=" * 70)
    print("EFFECT SIZES (Cohen's d) for Group Differences")
    print("=" * 70)

    for metric_name, metric_col in [('Efficiency', 'efficiency'),
                                      ('Lateral Deviation', 'mean_lateral_dev'),
                                      ('Heading Error', 'mean_heading_error'),
                                      ('Final Error', 'final_error')]:
        dv_vals = df_metrics[df_metrics['group'] == 'DV'][metric_col].dropna()
        ndv_vals = df_metrics[df_metrics['group'] == 'NDV'][metric_col].dropna()

        # Cohen's d
        pooled_std = np.sqrt(((len(dv_vals)-1)*dv_vals.std()**2 + (len(ndv_vals)-1)*ndv_vals.std()**2) /
                            (len(dv_vals) + len(ndv_vals) - 2))
        if pooled_std > 0:
            d = (dv_vals.mean() - ndv_vals.mean()) / pooled_std
        else:
            d = 0

        t_stat, p_val = stats.ttest_ind(dv_vals, ndv_vals)
        print(f"  {metric_name}: DV={dv_vals.mean():.3f}, NDV={ndv_vals.mean():.3f}, "
              f"d={d:.3f}, t={t_stat:.3f}, p={p_val:.4f}")

    # ── Effect sizes by Modality ──────────────────────────────────
    print("\n  Group differences (Cohen's d) by Modality:")
    for modality in MODALITY_ORDER:
        dv_eff = df_metrics[(df_metrics['group'] == 'DV') & (df_metrics['modality'] == modality)]['efficiency'].dropna()
        ndv_eff = df_metrics[(df_metrics['group'] == 'NDV') & (df_metrics['modality'] == modality)]['efficiency'].dropna()

        if len(dv_eff) > 0 and len(ndv_eff) > 0:
            pooled = np.sqrt(((len(dv_eff)-1)*dv_eff.std()**2 + (len(ndv_eff)-1)*ndv_eff.std()**2) /
                            (len(dv_eff) + len(ndv_eff) - 2))
            d = (dv_eff.mean() - ndv_eff.mean()) / pooled if pooled > 0 else 0
            t_stat, p_val = stats.ttest_ind(dv_eff, ndv_eff)
            print(f"    {MODALITY_SHORT[modality]}: DV={dv_eff.mean():.3f}, NDV={ndv_eff.mean():.3f}, "
                  f"d={d:.3f}, t={t_stat:.3f}, p={p_val:.4f}")

    print("\n" + "=" * 70)
    print("ANALYSIS COMPLETE")
    print("=" * 70)
    print(f"\nFigures saved to: {FIG_DIR}")
    print("  - fig9_trajectories_reconstructed.png")
    print("  - fig10_path_efficiency.png")
    print("  - fig11_lateral_deviation.png")
    print("  - fig12_heading_consistency.png")


if __name__ == '__main__':
    main()
